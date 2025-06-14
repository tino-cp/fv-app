import discord
from discord.ext import commands
from openpyxl import load_workbook

@commands.command(name="freeNumbers", aliases=["fr"])
async def free_numbers(ctx):
    # Load workbook and sheet
    wb = load_workbook("Formula V SuperLicense.xlsx", read_only=True, data_only=True)
    sheet = wb["Drivers"]

    unassigned = []

    # Loop through rows 8 to 106 (inclusive)
    for row in range(8, 107):
        number = sheet.cell(row=row, column=37).value  # AK
        name = sheet.cell(row=row, column=38).value    # AL

        if number is not None and (name is None or str(name).strip() == ""):
            unassigned.append(int(number))

    # Format the message
    if unassigned:
        unassigned.sort()
        numbers_str = ", ".join(str(n) for n in unassigned)
    else:
        numbers_str = "All numbers are assigned."

    # Send as embed
    embed = discord.Embed(
        title="🔢 These are the unassigned numbers:",
        description=numbers_str,
        color=discord.Color.blue()
    )
    await ctx.send(embed=embed)
