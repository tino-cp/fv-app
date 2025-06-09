import os
import datetime
import logging

import openpyxl
from discord.ext import tasks
from discord.utils  import get

USERS_FILE = "Users.xlsx"
GUILD_ID   = 843175947304960020  # ← make sure this is correct!

def ensure_users_workbook():
    """Create Users.xlsx with headers if it doesn’t already exist."""
    if not os.path.isfile(USERS_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["B1"] = "Date"
        ws["C1"] = "Time"
        ws["D1"] = "Member count"
        wb.save(USERS_FILE)

class MemberLogger:
    def __init__(self, bot):
        self.bot = bot
        ensure_users_workbook()
        self.log_member_count_daily.start()  # schedule the loop

    async def _get_guild(self):
        # first try the cache
        guild = get(self.bot.guilds, id=GUILD_ID)
        if guild is not None:
            return guild

        # fallback: fetch via HTTP
        try:
            guild = await self.bot.fetch_guild(GUILD_ID)
            return guild
        except Exception as e:
            logging.warning(f"[MemberLogger] Could not fetch guild {GUILD_ID}: {e}")
            return None

    async def _log_count(self):
        guild = await self._get_guild()
        if not guild:
            logging.warning(f"[MemberLogger] No guild object, skipping log.")
            return

        now      = datetime.datetime.now()
        date_str = now.strftime("%A, %-d %B %Y")  # e.g. "Monday, 9 June 2025"
        time_str = now.strftime("%H:%M")          # e.g. "23:10"
        count    = guild.member_count             # cached count

        wb = openpyxl.load_workbook(USERS_FILE)
        ws = wb.active
        ws.append([None, date_str, time_str, count])
        wb.save(USERS_FILE)

        logging.info(f"[MemberLogger] Logged {count} members at {date_str} {time_str}.")

    @tasks.loop(hours=24)
    async def log_member_count_daily(self):
        # this runs every 24h
        await self._log_count()

    @log_member_count_daily.before_loop
    async def before_log(self):
        # wait for ready, then do an immediate log
        await self.bot.wait_until_ready()
        await self._log_count()
